import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\ujaga\OneDrive\Desktop\Dashboard\ताल वाद्यपथक, पुणे -  नवीन सदस्य नोंदणी  -  गणेशोत्सव   २०२६ (Responses).xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb["Form Responses 1"]

print(f"Total rows in Excel: {ws.max_row}")

rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
print("Headers:", headers)
print("\n--- ALL RESPONSES ---")
for i, r in enumerate(rows[1:], 1):
    timestamp = r[0]
    name = r[1]
    phone = r[4]
    email = r[2]
    print(f"Row {i:2d}: [{timestamp}] | Name: {name} | Phone: {phone} | Email: {email}")
